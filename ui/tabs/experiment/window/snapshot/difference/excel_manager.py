import logging
import os

import pandas as pd
from PySide6.QtCore import QThread, Signal
from openpyxl.workbook import Workbook

from ui.tabs.experiment.window.snapshot.difference import ColorDifferenceModel
from util import local_storage_manager as lsm


class ExcelManager:
    def __init__(self, snapshot_path, plate_age, model):
        output_path = os.getenv("LOCAL_OUTPUT_PATH")
        self.snapshot_path = lsm.get_absolute_path(output_path, snapshot_path)
        self.plate_age = plate_age

        self.model = model
        self.targets = model.targets
        self.control_index = model.control_index
        self.control_name = model.targets.item_name(self.control_index)
        self.control_rgb_colors = model.get_rgb_colors(self.control_index)

    def get_path_to_save(self, target_index):
        file_name = f"{self.plate_age}H_{self.targets.item_name(target_index)}_{self.control_name}.xlsx"
        return os.path.join(self.snapshot_path, file_name)

    def save_target_colors(self, target_index):
        if target_index == self.control_index:
            return

        wb = Workbook()
        self._save_rgb_colors(wb, target_index)
        self._save_color_differences(wb, target_index, "xyy")
        self._save_color_differences(wb, target_index, "lab")
        wb.remove(wb["Sheet"])
        wb.save(self.get_path_to_save(target_index))

    def _save_rgb_colors(self, wb, target_index):
        target_rgb_colors, control_rgb_colors, rgb_differences = self.model.get_color_datas("rgb", target_index)
        target_name = self.targets.item_name(target_index)
        control_name = self.control_name

        rgb_sheet = wb.create_sheet("RGB Color")
        rgb_difference_sheet = wb.create_sheet("RGB difference")

        rgb_sheet.cell(1, 1, "Cell Info")
        rgb_sheet.cell(1, 2, f"{target_name} 'R'")
        rgb_sheet.cell(1, 3, f"{target_name} 'G'")
        rgb_sheet.cell(1, 4, f"{target_name} 'B'")
        rgb_sheet.cell(1, 5, f"{control_name} 'R'")
        rgb_sheet.cell(1, 6, f"{control_name} 'G'")
        rgb_sheet.cell(1, 7, f"{control_name} 'B'")

        rgb_difference_sheet.cell(1, 1, "Cell Info")
        rgb_difference_sheet.cell(1, 2, "RGB difference")

        for i, datas in enumerate(zip(target_rgb_colors, control_rgb_colors, rgb_differences)):
            solvent_i, additive_i = divmod(i, 8)
            cell_name = f"{chr(ord('A') + additive_i)}-{solvent_i + 1}"
            target, control, diff = datas[0], datas[1], datas[2]

            rgb_sheet.cell(i + 2, 1, cell_name)
            rgb_sheet.cell(i + 2, 2, target[0])
            rgb_sheet.cell(i + 2, 3, target[1])
            rgb_sheet.cell(i + 2, 4, target[2])
            rgb_sheet.cell(i + 2, 5, control[0])
            rgb_sheet.cell(i + 2, 6, control[1])
            rgb_sheet.cell(i + 2, 7, control[2])

            rgb_difference_sheet.cell(i + 2, 1, cell_name)
            rgb_difference_sheet.cell(i + 2, 2, diff)

    def _save_color_differences(self, wb, target_index, color_type):
        model: ColorDifferenceModel = self.model

        if color_type == "xyy":
            get_colors = model.get_xyy_colors
            sheet_name = "xyY difference"
        elif color_type == "lab":
            get_colors = model.get_lab_colors
            sheet_name = "Lab difference"
        else:
            get_colors = model.get_rgb_colors
            sheet_name = "RGB difference"

        target_colors = get_colors(target_index)
        control_colors = get_colors(self.control_index)
        color_differences = self.model.get_color_differences(target_colors, control_colors)

        sheet = wb.create_sheet(sheet_name)
        sheet.cell(1, 1, "Cell Info")
        sheet.cell(1, 2, sheet_name)

        for i, color_difference in enumerate(color_differences):
            solvent_i, additive_i = divmod(i, 8)
            cell_name = f"{chr(ord('A') + additive_i)}-{solvent_i + 1}"
            sheet.cell(i + 2, 1, cell_name)
            sheet.cell(i + 2, 2, color_difference)


class TimelineExcelWorker(QThread):
    finished = Signal(object, bool)

    def __init__(self, elapsed_times, rgb_datas, distance_datas, timeline_path, parent=None):
        super().__init__(parent)
        self.elapsed_times = elapsed_times
        self.rgb_datas = rgb_datas
        self.distance_datas = distance_datas
        self.timeline_path = timeline_path

    def run(self):
        try:
            TimelineExcelManager().save_timeline_datas(
                self.elapsed_times, self.rgb_datas, self.distance_datas, self.timeline_path
            )
            self.finished.emit(self, True)
        except Exception as e:
            logging.error(e)
            self.finished.emit(self, False)


class TimelineExcelManager:
    def save_timeline_datas(self, elapsed_times, rgb_datas, distance_datas, timeline_path):
        self.save_rgb_colors(elapsed_times, rgb_datas, timeline_path)
        self.save_distances(elapsed_times, distance_datas, timeline_path)

    def save_rgb_colors(self, elapsed_times, datas, timeline_path):
        output_path = os.getenv("LOCAL_OUTPUT_PATH")
        csv_path = lsm.get_absolute_path(output_path, f"{timeline_path}_rgb_colors.csv")

        datas.insert(0, "경과시간", elapsed_times)
        datas.to_csv(csv_path, index=False, encoding='utf-8-sig')

    def save_distances(self, elapsed_times, datas, timeline_path):
        output_path = os.getenv("LOCAL_OUTPUT_PATH")
        csv_path = lsm.get_absolute_path(output_path, f"{timeline_path}_distance_datas.csv")

        combined_data = []
        for col in datas.columns:
            combined_data.extend(zip(elapsed_times[1:], datas[col].iloc[1:]))

        df_combined = pd.DataFrame(combined_data, columns=["경과시간", "ColorDistance"])
        df_combined.to_csv(csv_path, index=False, encoding='utf-8-sig')
